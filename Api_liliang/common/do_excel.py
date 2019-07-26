#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @File  : do_excel.py
# @Author: LILIANG
# @Date  : 2019/6/3
# @Desc  :  test


import openpyxl
from common import logger
logger=logger.my_logger(__name__)

class Case:

    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
        self.sql = None

class DoExcel:

    def __init__(self,filename,sheetname):
        try:
            self.filename=filename
            self.sheetname=sheetname
            logger.info("加载excel文件")
            self.workbook=openpyxl.load_workbook(filename)
            logger.info("加载excel中的sheet文件")
            self.sheet=self.workbook[sheetname]
        except Exception as e:
            logger.info("目前测试环境为线上环境")('无法找到文件路径，报错{}'.format(e))

    #读取数据
    def get_cases(self):
        #获取最大行数
        max_row=self.sheet.max_row
        cases=[]
        for r in range(2,max_row+1):
            #实例化一个Case对象，通过对象调用属性,并赋值
            case = Case()  # 实例
            case.case_id = self.sheet.cell(row=r, column=1).value
            case.title = self.sheet.cell(row=r, column=2).value
            case.url = self.sheet.cell(row=r, column=3).value
            case.data = self.sheet.cell(row=r, column=4).value
            case.method = self.sheet.cell(row=r, column=5).value
            case.expected = self.sheet.cell(row=r, column=6).value
            case.sql = self.sheet.cell(row=r, column=9).value  # sql
            cases.append(case)

        self.workbook.close()
        return cases


    #数据回写

    def write_result(self,row,actual, result):
        sheet=self.workbook[self.sheetname]
        sheet.cell(row,7).value=actual
        sheet.cell(row,8).value=result
        self.workbook.save(filename=self.filename)
        self.workbook.close()


if __name__ == '__main__':
    from common import http_requests,logger
    from common import constants
    logger=logger.my_logger('case')
    logger.info('读取excel')
    # do_excel = DoExcel(r"..\data\xiaohai0412_testcases.xlsx", sheetname='login')
    #从配置文件中读取文件的路径
    do_excel = DoExcel(constants.case_file, sheetname='login')
    cases = do_excel.get_cases()
    http_request = http_requests.HttpRequests1()
    for case in cases:
        # print(case.case_id)
        # print(case.method)
        # print(case.data)
        #将对象case的数据以字典的形式打印出来
        print(case.__dict__)
        print(type(case.data))

        # case.data=eval(case.data)
        #这里的类型转换（str=>dict)可以优化到封装http_requests层次做优化
        resp = http_request.requests(case.method, case.url, case.data)
        print(resp.status_code)
        print(resp.text)  # 响应文本
        resp_dict = resp.json()  # 返回字典
        print(resp_dict)

        actual = resp.text
        if case.expected == actual:  # 判断期望结果是否与实际结果一致
            do_excel.write_result(case.case_id + 1, actual, 'PASS')

        else:
            do_excel.write_result(case.case_id + 1, actual, 'FAIL')
